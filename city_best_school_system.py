import mysql.connector

# Replace these values with your actual MySQL username and password
mysql_username = 'israelaidoo'  # Your MySQL username
mysql_password = 'Kobbie@2025.'  # Your MySQL password
mysql_username = "israelaidoo"
mysql_password = "Kobbie@2025."

script = f"""

import mysql.connector
from datetime import date

# Connect to MySQL
def connect():
    return mysql.connector.connect(
        host="localhost",
        user="israelaidoo",
        password="Kobbie@2025.",
        database="citybest_school"
    )

# Add a new student
def add_student():
    first_name = input("First Name: ")
    last_name = input("Last Name: ")
    dob = input("Date of Birth (YYYY-MM-DD): ")
    gender = input("Gender (Male/Female): ")
    class_id = int(input("Class ID: "))
    guardian = input("Guardian Name: ")
    contact = input("Contact Number: ")
    address = input("Address: ")

    conn = connect()
    cursor = conn.cursor()
    query = \"\"\"
        INSERT INTO students (first_name, last_name, date_of_birth, gender, class_id,
                              guardian_name, contact_number, address, enrollment_date)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    \"\"\"
    values = (first_name, last_name, dob, gender, class_id, guardian, contact, address, date.today())
    cursor.execute(query, values)
    conn.commit()
    print("Student added successfully.")
    cursor.close()
    conn.close()

# View all students
def view_students():
    conn = connect()
    cursor = conn.cursor()
    cursor.execute("SELECT student_id, first_name, last_name, class_id FROM students")
    rows = cursor.fetchall()
    for row in rows:
        print(row)
    cursor.close()
    conn.close()

# Add a fee record
def add_fee():
    student_id = int(input("Student ID: "))
    term = input("Term (e.g. Term 1): ")
    amount = float(input("Amount Paid: "))
    status = input("Status (Paid/Partial/Unpaid): ")

    conn = connect()
    cursor = conn.cursor()
    query = \"\"\"
        INSERT INTO fees (student_id, term, amount, date_paid, status)
        VALUES (%s, %s, %s, %s, %s)
    \"\"\"
    values = (student_id, term, amount, date.today(), status)
    cursor.execute(query, values)
    conn.commit()
    print("Fee record added.")
    cursor.close()
    conn.close()

# View unpaid fees
def view_unpaid_fees():
    conn = connect()
    cursor = conn.cursor()
    cursor.execute(\"\"\"
        SELECT students.first_name, students.last_name, fees.term, fees.amount
        FROM fees
        JOIN students ON students.student_id = fees.student_id
        WHERE fees.status != 'Paid'
    \"\"\")
    rows = cursor.fetchall()
    for row in rows:
        print(row)
    cursor.close()
    conn.close()

# Record attendance
def record_attendance():
    student_id = int(input("Student ID: "))
    status = input("Status (Present/Absent): ")

    conn = connect()
    cursor = conn.cursor()
    query = \"\"\"
        INSERT INTO attendance (student_id, date, status)
        VALUES (%s, %s, %s)
    \"\"\"
    values = (student_id, date.today(), status)
    cursor.execute(query, values)
    conn.commit()
    print("Attendance recorded.")
    cursor.close()
    conn.close()

# View attendance
def view_attendance():
    student_id = int(input("Student ID: "))

    conn = connect()
    cursor = conn.cursor()
    cursor.execute(\"\"\"
        SELECT date, status
        FROM attendance
        WHERE student_id = %s
        ORDER BY date DESC
    \"\"\", (student_id,))
    rows = cursor.fetchall()
    for row in rows:
        print(row)
    cursor.close()
    conn.close()

# Add performance
def add_performance():
    student_id = int(input("Student ID: "))
    subject = input("Subject: ")
    term = input("Term: ")
    score = float(input("Score: "))

    conn = connect()
    cursor = conn.cursor()
    query = \"\"\"
        INSERT INTO performance (student_id, subject_name, term, score)
        VALUES (%s, %s, %s, %s)
    \"\"\"
    cursor.execute(query, (student_id, subject, term, score))
    conn.commit()
    print("Performance added.")
    cursor.close()
    conn.close()

# View report
def generate_report():
    student_id = int(input("Student ID: "))
    conn = connect()
    cursor = conn.cursor()

    print("=== STUDENT PERFORMANCE REPORT ===")
    cursor.execute("SELECT first_name, last_name FROM students WHERE student_id = %s", (student_id,))
    student = cursor.fetchone()
    if student:
        print(f"Name: {student[0]} {student[1]}")

        cursor.execute("SELECT subject_name, term, score FROM performance WHERE student_id = %s", (student_id,))
        scores = cursor.fetchall()
        if scores:
            print("Subject | Term | Score")
            for sub in scores:
                print(f"{sub[0]} | {sub[1]} | {sub[2]}")
        else:
            print("No performance records found.")
    else:
        print("Student not found.")
    cursor.close()
    conn.close()

# Menu
def menu():
    while True:
        print(\"""
        === CITYBEST SCHOOL DATABASE MENU ===
        1. Add Student
        2. View Students
        3. Add Fee Record
        4. View Unpaid Fees
        5. Record Attendance
        6. View Attendance
        7. Add Performance
        8. Generate Report
        9. Exit
        \""")
        choice = input("Select an option (1-9): ")

        if choice == '1':
            add_student()
        elif choice == '2':
            view_students()
        elif choice == '3':
            add_fee()
        elif choice == '4':
            view_unpaid_fees()

        elif choice == '5':
            record_attendance()
        elif choice == '6':
            view_attendance()
        elif choice == '7':
            add_performance()
        elif choice == '8':
            generate_report()
        elif choice == '9':
            print("Goodbye!")
            break
        else:
            print("Invalid option. Try again.")

if __name__ == "__main__":
    menu()
"""

Script




# citybest_school_system/main.py
import mysql.connector
import hashlib
import os
from datetime import date
from getpass import getpass
from openpyxl import Workbook
from fpdf import FPDF

# Constants for file paths
EXCEL_DIR = "/mnt/data/citybest_school_system/excel_reports"
PDF_DIR = "/mnt/data/citybest_school_system/pdf_reports"

# Ensure directories exist
os.makedirs(EXCEL_DIR, exist_ok=True)
os.makedirs(PDF_DIR, exist_ok=True)

# Connect to MySQL
def connect():
    return mysql.connector.connect(
        host="localhost",
        user="israelaidoo",
        password="Kobbie@2025.",
        database="citybest_school"
    )

# Hash password
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# User login
def login():
    username = input("Username: ")
    password = getpass("Password: ")
    hashed = hash_password(password)

    conn = connect()
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE username=%s AND password=%s", (username, hashed))
    result = cursor.fetchone()
    cursor.close()
    conn.close()

    if result:
        print(f"Welcome {username} ({result[0]})")
        return result[0]  # Return role
    else:
        print("Invalid login.")
        return None

# Add new user (Admin only)
def add_user():
    username = input("New Username: ")
    password = getpass("Password: ")
    role = input("Role (admin/teacher): ").lower()
    hashed = hash_password(password)

    conn = connect()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO users (username, password, role) VALUES (%s, %s, %s)", (username, hashed, role))
    conn.commit()
    print("User added successfully.")
    cursor.close()
    conn.close()

# Add Student
def add_student():
    conn = connect()
    cursor = conn.cursor()
    fname = input("First name: ")
    lname = input("Last name: ")
    dob = input("DOB (YYYY-MM-DD): ")
    gender = input("Gender: ")
    class_id = input("Class ID: ")
    guardian = input("Guardian name: ")
    contact = input("Contact number: ")
    address = input("Address: ")
    today = date.today()
    cursor.execute("""
        INSERT INTO students (first_name, last_name, date_of_birth, gender, class_id,
                              guardian_name, contact_number, address, enrollment_date)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (fname, lname, dob, gender, class_id, guardian, contact, address, today))
    conn.commit()
    print("Student added.")
    cursor.close()
    conn.close()

# Generate Excel Report
def export_students_to_excel():
    conn = connect()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM students")
    data = cursor.fetchall()
    headers = [i[0] for i in cursor.description]

    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in data:
        ws.append(row)

    file_path = os.path.join(EXCEL_DIR, f"students_{date.today()}.xlsx")
    wb.save(file_path)
    print(f"Excel report saved to {file_path}")
    cursor.close()
    conn.close()

# Generate PDF Report
def export_students_to_pdf():
    conn = connect()
    cursor = conn.cursor()
    cursor.execute("SELECT student_id, first_name, last_name, class_id FROM students")
    data = cursor.fetchall()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="CityBest Students Report", ln=True, align="C")

    for row in data:
        pdf.cell(200, 10, txt=f"ID: {row[0]} - {row[1]} {row[2]} (Class {row[3]})", ln=True)

    file_path = os.path.join(PDF_DIR, f"students_{date.today()}.pdf")
    pdf.output(file_path)
    print(f"PDF report saved to {file_path}")
    cursor.close()
    conn.close()

# Main Menu
def menu(role):
    while True:
        print("""
        --- CITYBEST SCHOOL SYSTEM ---
        1. Add Student
        2. Export Students (Excel)
        3. Export Students (PDF)
        4. Add User (Admin Only)
        5. Logout
        """)
        choice = input("Choose: ")

        if choice == '1' and role in ['admin', 'teacher']:
            add_student()
        elif choice == '2' and role in ['admin', 'teacher']:
            export_students_to_excel()
        elif choice == '3' and role in ['admin', 'teacher']:
            export_students_to_pdf()
        elif choice == '4' and role == 'admin':
            add_user()
        elif choice == '5':
            break
        else:

            print("Invalid choice or insufficient permissions.")

# Start app
if __name__ == "__main__":
    role = login()
    if role:
        menu(role)

